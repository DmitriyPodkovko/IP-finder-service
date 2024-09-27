import os
from typing import Any
import openpyxl
import logging
from config.settings import (EXCEL_ROW_COLUMN,
                             DESTINATION_NUMBER,
                             EXCEL_OUTPUT_FILE_PREFIX)


class ExcelHandler:
    def __init__(self, xlsx_file) -> None:
        self._xlsx_file = xlsx_file
        self._DST_column_idx = None
        self._xlsx_output_file = None
        self._current_row = EXCEL_ROW_COLUMN.get('default').get('Start_row')
        self.errors = ''

    def get_ip_list_from_xlsx_file(self) -> list[tuple[Any, Any, Any, Any, Any]] | list[Any]:
        try:
            ip_list = []
            workbook = openpyxl.load_workbook(self._xlsx_file)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=EXCEL_ROW_COLUMN.get('default').get('Start_row'), values_only=True):
                ip_tuple = (
                    row[EXCEL_ROW_COLUMN.get('default').get('IP_DST')],
                    row[EXCEL_ROW_COLUMN.get('default').get('Port_DST')],
                    row[EXCEL_ROW_COLUMN.get('default').get('Date')],
                    row[EXCEL_ROW_COLUMN.get('default').get('Time')],
                    row[EXCEL_ROW_COLUMN.get('default').get('Provider')]
                )
                # Check if all elements in the tuple are not empty
                if all(ip_tuple):
                    ip_list.append(ip_tuple)
                    print(f'IP_DST, Port_DST, Date, Time, Provider: {ip_tuple}')
                    logging.info(f'IP_DST, Port_DST, Date, Time, Provider: {ip_tuple}')
            workbook.close()
            return ip_list
        except Exception as e:
            print(f'Error input .xlsx file:\n {self._xlsx_file} ({str(e)})')
            logging.error(f'Error input .xlsx file:\n {self._xlsx_file} ({str(e)})')
            self.errors += 'get xlsx error\n'
            return []

    def create_output_xlsx_file(self, user_directory) -> None:
        try:
            workbook = openpyxl.load_workbook(self._xlsx_file)
            sheet = workbook.active
            if EXCEL_ROW_COLUMN.get('default').get('DST') >= 0:
                self._DST_column_idx = EXCEL_ROW_COLUMN.get('default').get('DST') + 1
            else:
                self._DST_column_idx = sheet.max_column + 1
                sheet.cell(row=1, column=self._DST_column_idx, value=DESTINATION_NUMBER)
            file_name, file_extensions = os.path.splitext(self._xlsx_file)
            file_name = os.path.basename(file_name)
            self._xlsx_output_file = f'{user_directory}/{file_name}{EXCEL_OUTPUT_FILE_PREFIX}.xlsx'
            workbook.save(self._xlsx_output_file)
            workbook.close()
            print(f'CREATED: {self._xlsx_output_file}')
            logging.info(f'CREATED: {self._xlsx_output_file}')
        except Exception as e:
            print(f'Error creating .xlsx file:\n {self._xlsx_file} ({str(e)})')
            logging.error(f'Error creating .xlsx file:\n {self._xlsx_file} ({str(e)})')
            self.errors += 'create xlsx error\n'

    def save_result_to_output_xlsx_file(self, dst_list: list) -> None:
        try:
            workbook = openpyxl.load_workbook(self._xlsx_output_file)
            sheet = workbook.active
            for dst_set in dst_list:
                current_column_idx = self._DST_column_idx
                for number in dst_set:
                    sheet.cell(row=self._current_row, column=current_column_idx, value=number)
                    current_column_idx += 1
                self._current_row += 1
            workbook.save(self._xlsx_output_file)
            workbook.close()
            print(f'saved')
            logging.info(f'saved')
        except Exception as e:
            print(f'Error saving .xlsx:\n {self._xlsx_file} ({str(e)})')
            logging.error(f'Error saving .xlsx:\n {self._xlsx_file} ({str(e)})')
            self.errors += 'save xlsx error\n'
